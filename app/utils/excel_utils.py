"""
Excel处理相关工具函数
"""

import os
from datetime import datetime, date
from typing import List, Dict, Any, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.config import settings
from utils.file_utils import get_unique_filename


def create_excel_file(data: List[Dict[str, Any]], filename: str = None) -> str:
    """
    创建Excel文件
    
    Args:
        data: 数据列表，每个元素是一个字典
        filename: 文件名，如果不提供则自动生成
        
    Returns:
        创建的Excel文件路径
    """
    if not data:
        raise ValueError("数据不能为空")
    
    # 如果没有提供文件名，则生成一个
    if not filename:
        filename = get_unique_filename("report.xlsx")
    
    # 确保文件扩展名是.xlsx
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    
    # 确保目录存在
    file_path = os.path.join(settings.UPLOAD_DIR, filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # 使用pandas创建DataFrame并保存为Excel
    df = pd.DataFrame(data)
    df.to_excel(file_path, index=False, engine='openpyxl')
    
    return file_path


def create_attendance_report_excel(
    attendance_data: List[Dict[str, Any]],
    month: str,
    filename: str = None
) -> str:
    """
    创建考勤报告Excel文件
    
    Args:
        attendance_data: 考勤数据列表
        month: 报告月份
        filename: 文件名，如果不提供则自动生成
        
    Returns:
        创建的Excel文件路径
    """
    if not filename:
        filename = get_unique_filename(f"attendance_report_{month}.xlsx")
    
    # 确保文件扩展名是.xlsx
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    
    # 确保目录存在
    file_path = os.path.join(settings.UPLOAD_DIR, filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}考勤报告"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题
    title = f"{month}考勤报告"
    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入表头
    headers = ['员工编号', '姓名', '部门', '日期', '上班时间', '下班时间', '工作时长', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row, record in enumerate(attendance_data, 4):
        ws.cell(row=row, column=1, value=record.get('employee_id', ''))
        ws.cell(row=row, column=2, value=record.get('employee_name', ''))
        ws.cell(row=row, column=3, value=record.get('department_name', ''))
        ws.cell(row=row, column=4, value=record.get('date', ''))
        ws.cell(row=row, column=5, value=record.get('check_in_time', ''))
        ws.cell(row=row, column=6, value=record.get('check_out_time', ''))
        ws.cell(row=row, column=7, value=record.get('work_hours', ''))
        ws.cell(row=row, column=8, value=record.get('status', ''))
        
        # 添加边框
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border
    
    # 调整列宽
    column_widths = [12, 12, 15, 12, 12, 12, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 保存文件
    wb.save(file_path)
    
    return file_path


def create_leave_report_excel(
    leave_data: List[Dict[str, Any]],
    month: str,
    filename: str = None
) -> str:
    """
    创建请假报告Excel文件
    
    Args:
        leave_data: 请假数据列表
        month: 报告月份
        filename: 文件名，如果不提供则自动生成
        
    Returns:
        创建的Excel文件路径
    """
    if not filename:
        filename = get_unique_filename(f"leave_report_{month}.xlsx")
    
    # 确保文件扩展名是.xlsx
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    
    # 确保目录存在
    file_path = os.path.join(settings.UPLOAD_DIR, filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}请假报告"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题
    title = f"{month}请假报告"
    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入表头
    headers = ['员工编号', '姓名', '部门', '请假类型', '开始日期', '结束日期', '状态']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for row, record in enumerate(leave_data, 4):
        ws.cell(row=row, column=1, value=record.get('employee_id', ''))
        ws.cell(row=row, column=2, value=record.get('employee_name', ''))
        ws.cell(row=row, column=3, value=record.get('department_name', ''))
        ws.cell(row=row, column=4, value=record.get('leave_type', ''))
        ws.cell(row=row, column=5, value=record.get('start_date', ''))
        ws.cell(row=row, column=6, value=record.get('end_date', ''))
        ws.cell(row=row, column=7, value=record.get('status', ''))
        
        # 添加边框
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
    
    # 调整列宽
    column_widths = [12, 12, 15, 12, 12, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 保存文件
    wb.save(file_path)
    
    return file_path


def create_statistics_report_excel(
    statistics_data: Dict[str, Any],
    month: str,
    filename: str = None
) -> str:
    """
    创建统计报告Excel文件
    
    Args:
        statistics_data: 统计数据字典
        month: 报告月份
        filename: 文件名，如果不提供则自动生成
        
    Returns:
        创建的Excel文件路径
    """
    if not filename:
        filename = get_unique_filename(f"statistics_report_{month}.xlsx")
    
    # 确保文件扩展名是.xlsx
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    
    # 确保目录存在
    file_path = os.path.join(settings.UPLOAD_DIR, filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{month}统计报告"
    
    # 定义样式
    title_font = Font(name='微软雅黑', size=16, bold=True)
    header_font = Font(name='微软雅黑', size=12, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题
    title = f"{month}考勤统计报告"
    ws.merge_cells('A1:D1')
    ws['A1'] = title
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    
    # 写入总体统计
    ws['A3'] = "总体统计"
    ws['A3'].font = header_font
    
    overall_stats = statistics_data.get('overall', {})
    ws['A4'] = "总员工数"
    ws['B4'] = overall_stats.get('total_employees', 0)
    
    ws['A5'] = "正常出勤天数"
    ws['B5'] = overall_stats.get('normal_days', 0)
    
    ws['A6'] = "迟到次数"
    ws['B6'] = overall_stats.get('late_count', 0)
    
    ws['A7'] = "早退次数"
    ws['B7'] = overall_stats.get('early_leave_count', 0)
    
    ws['A8'] = "缺勤次数"
    ws['B8'] = overall_stats.get('absence_count', 0)
    
    ws['A9'] = "请假天数"
    ws['B9'] = overall_stats.get('leave_days', 0)
    
    # 写入部门统计
    ws['A11'] = "部门统计"
    ws['A11'].font = header_font
    
    ws['A12'] = "部门名称"
    ws['B12'] = "员工数"
    ws['C12'] = "出勤率"
    ws['D12'] = "请假率"
    
    department_stats = statistics_data.get('departments', [])
    for row, dept in enumerate(department_stats, 13):
        ws.cell(row=row, column=1, value=dept.get('name', ''))
        ws.cell(row=row, column=2, value=dept.get('employee_count', 0))
        ws.cell(row=row, column=3, value=dept.get('attendance_rate', 0))
        ws.cell(row=row, column=4, value=dept.get('leave_rate', 0))
        
        # 添加边框
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
    
    # 调整列宽
    column_widths = [15, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 保存文件
    wb.save(file_path)
    
    return file_path


def read_excel_file(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """
    读取Excel文件
    
    Args:
        file_path: 文件路径
        sheet_name: 工作表名称，如果不提供则读取第一个工作表
        
    Returns:
        DataFrame对象
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    if sheet_name:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        return pd.read_excel(file_path)


def export_to_excel(
    data: Union[List[Dict[str, Any]], pd.DataFrame],
    filename: str,
    sheet_name: str = "Sheet1"
) -> str:
    """
    导出数据到Excel文件
    
    Args:
        data: 数据，可以是字典列表或DataFrame
        filename: 文件名
        sheet_name: 工作表名称
        
    Returns:
        保存的文件路径
    """
    # 确保文件扩展名是.xlsx
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"
    
    # 确保目录存在
    file_path = os.path.join(settings.UPLOAD_DIR, filename)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    # 如果是字典列表，转换为DataFrame
    if isinstance(data, list):
        df = pd.DataFrame(data)
    else:
        df = data
    
    # 保存到Excel
    df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
    
    return file_path


def append_to_excel(
    data: Union[List[Dict[str, Any]], pd.DataFrame],
    file_path: str,
    sheet_name: str = None
) -> None:
    """
    追加数据到Excel文件
    
    Args:
        data: 数据，可以是字典列表或DataFrame
        file_path: 文件路径
        sheet_name: 工作表名称，如果不提供则使用第一个工作表
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    # 如果是字典列表，转换为DataFrame
    if isinstance(data, list):
        df = pd.DataFrame(data)
    else:
        df = data
    
    # 读取现有数据
    if sheet_name:
        existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
    else:
        existing_df = pd.read_excel(file_path)
    
    # 合并数据
    combined_df = pd.concat([existing_df, df], ignore_index=True)
    
    # 保存回文件
    combined_df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')